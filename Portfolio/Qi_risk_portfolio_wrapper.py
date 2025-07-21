import os
from abc import ABC, abstractmethod
from typing import Callable, Dict, List, Optional, Tuple

from datetime import datetime

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import openpyxl
from dateutil.relativedelta import relativedelta
from matplotlib.ticker import PercentFormatter
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


import numpy as np
import pandas as pd
import qi_client
from joblib import Parallel, delayed
from qi_client.rest import ApiException

QI_API_KEY = os.environ.get('QI_API_KEY', None)
if not QI_API_KEY:
    print("Mandatory environment variable QI_API_KEY not set!")

configuration = qi_client.Configuration()
configuration.api_key['X-API-KEY'] = QI_API_KEY
api_instance = qi_client.DefaultApi(qi_client.ApiClient(configuration))

RISK_WIN = 125
HALF_LIFE = 90
ZERO_IDX = 0



COLORS = {
    '10Y Yield': '#a6cee3',
    'USD 10Y Yield': '#a6cee3',
    'US 10Y Yield': '#a6cee3',
    'EU 10Y Yield': '#a6cee3',
    'DM Yield Ccy1': '#a6cee3',
    'CB QT Expectations': '#1f78b4',
    'FED QT Expectations': '#1f78b4',
    'ECB QT Expectations': '#1f78b4',
    'CB QT Expectations Ccy1': '#1f78b4',
    'CB QT Expectations Ccy2': '#1f78b4',
    'CB Rate Expectations': '#b2df8a',
    'FED Rate Expectations': '#b2df8a',
    'ECB Rate Expectations': '#b2df8a',
    'CB Rate Expectations Ccy1': '#b2df8a',
    'CB Rate Expectations Ccy2': '#b2df8a',
    'Corporate Credit': '#33a02c',
    'US Corporate Credit': '#33a02c',
    'EU Corporate Credit': '#33a02c',
    'US HY': '#33a02c',
    'DM FX': '#fb9a99',
    'EM FX': '#fb9a99',
    'USD TWI': '#fb9a99',
    'Economic Growth': '#e31a1c',
    'Growth Expec. Diff.': '#e31a1c',
    'China Stress': '#e31a1c',
    'China Growth': '#e31a1c',
    'US Growth': '#e31a1c',
    'EU Growth': '#e31a1c',
    'US GDP': '#e31a1c',
    'Energy': '#fdbf6f',
    'WTI': '#fdbf6f',
    'Forward Growth Expectations': '#ff7f00',
    'US Forward Growth Expectations': '#ff7f00',
    'EU Forward Growth Expectations': '#ff7f00',
    'US 5s30s Swap': '#ff7f00',
    'Inflation': '#cab2d6',
    'EU Inflation': '#cab2d6',
    'US Inflation': '#cab2d6',
    'US 5Y Infl. Expec.': '#cab2d6',
    'Market': 'black',
    'Metals': '#6a3d9a',
    'Copper': '#6a3d9a',
    'Real Rates': '#ffff99',
    'Rate Diff.': '#ffff99',
    'ECB Real Rates': '#ffff99',
    'US Real Rates': '#ffff99',
    'EU Real Rates': '#ffff99',
    'USD 10Y Real Rate': '#ffff99',
    'Risk Aversion': '#b15928',
    'VIX': '#b15928',
    "total": "#1F77B4",
    "total_risk": "#1F77B4",
    "factor_risk": "#D62728",
    "factor": "#D62728",
    "specific_risk": "#2CA02C",
    "specific": "#2CA02C",
    "actual": "#1F77B4",
    "factor": "#D62728",
    "specific": "#2CA02C",
    "predicted_vol": "#1F77B4",
    "realised_vol": "#D62728",
    "Idio": "#483D8B",
    "SPX Index": "#8BA4FA",
    "actual": "#8BA4FA",
    'factor': '#D62728',
    "specific": "#72CC50",
    'Financial Conditions': '#8BA4FA',
    'Economic Fundamentals': '#FFCDEA',
    'Risk Appetite': '#05C2ED',
    "Macro": "#1F77B4",
    '1month_macro_level_change': "#1F77B4",
    'today': '#ffff99',
    '1w': '#a6cee3',
    '1m': '#1f78b4',
    '3m': '#e31a1c',
    '6m': '#6a3d9a',
    '12m': '#33a02c',
    '-1w': '#a6cee3',
    '-1m': '#1f78b4',
    '-3m': '#e31a1c',
    '-6m': '#6a3d9a',
    '-12m': '#33a02c',
}

class PorfolioRiskExcel():
    @staticmethod
    def add_description_sheet(workbook, sheet_name="Description"):
        """
        Adds a description sheet to the given workbook with predefined data and formatting.

        Args:
            workbook (Workbook): The openpyxl Workbook object to which the sheet will be added.
            data (dict): A dictionary with "Sheet" and "Description" keys containing the data.
            sheet_name (str): Name of the sheet to be added.
        """
        ws = workbook.create_sheet(title=sheet_name)

        data = {
            "Sheet": [
                "charts",
                "weights",
                "exposures",
                "factor_attribution",
                "risk_and_return_ts",
                "factor_risk_proportion",
                "factor_risk_contribution",
                "exposures_yyyy-mm-dd",
                "factor_attribution_3m",
                "single_stock_risk_yyyy-mm-dd",
                "port_stock_risk_yyyy-mm-dd",
                "port_stock_MCTR_yyyy-mm-dd",
                "port_stock_prop_yyyy-mm-dd",
                "stock_risk_proportion",
                "stock_risk_contribution",
                "Factor Glossary",
            ],
            "Description": [
                "Example charts using data",
                "Fixed weights (sums to 1 if long only)\nLong or Short",
                "Weighted exposure to factors  (%)  per factor daily standard deviation move (250d).  (values expressed in % eg 2 =  2%)",
                "Daily portfolio % return attributable to each individual factor - each day sums to total daily factor return.  (values expressed in % eg 2 =  2%)",
                "total_return - actual daily portfolio % return  (values expressed in % eg 2 =  2%)\nfactor_return - daily portfolio % return attributable to all factors i.e. total factor daily % return\nspecific_return - daily portfolio % return NOT explained by factors i.e. idiosyncratic\ntotal_risk (vol %) - daily portfolio % predicted total risk (multiply by sqrt(252) to annualise)\nfactor_risk (vol %) - daily portfolio % predicted factor risk (factor risk^2 + specific risk^2 = total risk^2)\nspecific_risk (vol %) - daily portfolio % predicted specific risk (factor risk^2 + specific risk^2 = total risk^2)",
                "% of total portfolio risk attributable to each individual factor & specific, each day (sums to 100%)",
                "Daily % predicted risk attributable to each individual factor & specific, which linearly sums to daily portfolio % predicted total risk",
                "% factor exposure of each individual security within the portfolio on stated date",
                "Fixed 3mth % portfolio return attributable to each individual security for each factor",
                "Portfolio's predicted risk attributable to each security for each factor on stated date in Vol %, assuming each individual security is analyzed in isolation",
                "Portfolio's predicted risk attributable to each security for each factor on stated date in Vol % - securities are not analyzed in isolation",
                "Portfolio's predicted risk attributable to each security for each factor on stated date in MCTR %; sums linearly to total risk  - securities are not analyzed in isolation",
                "% of total portfolio risk attributable to each security for each factor on stated date; sums to 1.0 ",
                "% of total portfolio risk explained by each individual security for each factor and specific, each day; sums to 1.0",
                "Daily portfolio % predicted risk by factor & specific attributable to each individual security; sums linearly to total risk",
                "Factor definitions"
            ],
        }

        # Add headers to specific cells
        ws["B2"] = "Sheet"
        ws["E2"] = "Description"

        # Style headers
        header_font = Font(bold=True)
        ws["B2"].font = header_font
        ws["E2"].font = header_font

        # Write the data starting from row 4, with an empty row between data rows
        start_row = 4
        row_increment = 2  # Leave one empty row between data rows
        current_row = start_row

        for sheet_name, description in zip(data["Sheet"], data["Description"]):
            # Write data
            ws[f"B{current_row}"] = sheet_name
            ws[f"E{current_row}"] = description

            # Bold the cell in column B
            ws[f"B{current_row}"].font = Font(bold=True)

            current_row += row_increment  # Skip one row

        # Adjust column widths for readability
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["E"].width = 100

        # Define a bottom border style
        thin_bottom_border = Border(bottom=Side(style="thin"))

        # Apply bottom border to header row and last row of data
        for row in [2, current_row - row_increment]:
            for col in range(2, 6):  # Columns B (2) to E (5)
                ws.cell(row=row, column=col).border = thin_bottom_border

        # Hide gridlines (optional)
        ws.sheet_view.showGridLines = False

    @staticmethod
    def add_disclaimer_sheet(workbook, sheet_name="Disclaimer"):
        """
        Adds a disclaimer sheet to the workbook with formatted text starting in cell B4.

        Args:
            workbook (Workbook): The openpyxl Workbook object to which the sheet will be added.
            disclaimer_text (str): The text of the disclaimer to include in the sheet.
            sheet_name (str): Name of the sheet to be added.
        """
        # Create a new sheet for the disclaimer
        ws = workbook.create_sheet(title=sheet_name)

        disclaimer_text = (
            "This document is being sent only to investment professionals (as that term is defined in article 19(5) of the Financial Services "
            "and Markets Act 2000 (Financial Promotion) OrderSq005 ('FPO')) or to persons to whom it would otherwise be lawful to distribute it. "
            "Accordingly, persons who do not have professional experience in matters relating to investments should not rely on this document. "
            "The information contained herein is for general guidance and information only and is subject to amendment or correction. This document "
            "is not directed to, or intended for distribution to or use by, any person or entity who is a citizen or resident of or located in any locality, "
            "state, country or other jurisdiction where such distribution, publication, availability or use would be contrary to law or regulation.\n\n"
            "This document is provided for information purposes only, is intended for your use only, and does not constitute an invitation or offer to "
            "subscribe for or purchase any securities, any product or any service and neither this document nor anything contained herein shall form the "
            "basis of any contract or commitment whatsoever. This document does not constitute any recommendation regarding any securities, futures, "
            "derivatives or other investment products. The information contained herein is provided for informational and discussion purposes only and is "
            "not and, may not be relied on in any manner as accounting, legal, tax, investment, regulatory or other advice.\n\n"
            "Information and opinions presented in this document have been obtained or derived from sources believed to be reliable, but Quant Insight Limited "
            "(Qi) makes no representation as to their accuracy or completeness or reliability and expressly disclaims any liability, including incidental or "
            "consequential damages arising from errors in this publication. No reliance may be placed for any purpose on the information and opinions contained "
            "in this document. No representation, warranty or undertaking, express or implied, is given as to the accuracy or completeness of the information "
            "or opinions contained in this document by any of Qi, its employees or affiliates and no liability is accepted by such persons for the accuracy or "
            "completeness of any such information or opinions. Any data provided in this document indicating past performance is not a reliable indicator of "
            "future returns/performance. Nothing contained herein shall be relied upon as a promise or representation whether as to past or future performance.\n\n"
            "This presentation is strictly confidential and may not be reproduced or redistributed in whole or in part nor may its contents be disclosed to any "
            "other person under any circumstances without the express permission of Quant Insight Limited."
        )


        # Add the title "Disclaimer" in cell A2
        ws["A2"] = "Disclaimer"
        ws["A2"].font = Font(bold=True, size=14)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="top")

        # Add the disclaimer text starting from cell B4
        ws["B4"] = disclaimer_text
        ws["B4"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws["B4"].font = Font(size=11)

        # Merge cells for the disclaimer text if needed (optional)
        ws.merge_cells("B4:K20")  # Adjust range as needed for your text size

        # Apply a thin border around the disclaimer text area
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=4, max_row=20, min_col=2, max_col=11):
            for cell in row:
                cell.border = thin_border

        # Adjust column widths for readability
        for col in range(2, 12):  # Columns B to K
            ws.column_dimensions[chr(64 + col)].width = 12

        # Set the row height for better readability
        for row in range(1, 21):  # Rows 1 to 20
            ws.row_dimensions[row].height = 20

        # Hide gridlines (optional)
        ws.sheet_view.showGridLines = False


    @staticmethod
    def add_factor_glossary_sheet(workbook, df, sheet_name="Factor Glossary"):
        """
        Adds a Factor Glossary sheet to the workbook based on a given DataFrame,
        with the table starting in cell B4 and no gridlines.

        Args:
            workbook (Workbook): The openpyxl Workbook object to which the sheet will be added.
            df (pd.DataFrame): The DataFrame containing the Factor Glossary data.
            sheet_name (str): Name of the sheet to be added.
        """
        # Create a new sheet
        ws = workbook.create_sheet(title=sheet_name)

        # Hide gridlines
        ws.sheet_view.showGridLines = False

        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="6ABDE9", end_color="6ABDE9", fill_type="solid")
        text_font = Font(size=11)
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers, starting in cell B4
        start_row = 4
        start_col = 2  # Column B
        for col_num, column_title in enumerate(df.columns, start=start_col):
            cell = ws.cell(row=start_row, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # Write data rows, starting below the headers
        for row_num, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_num, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = text_font
                cell.alignment = alignment
                cell.border = border

        # Adjust column widths
        for col_num, column_title in enumerate(df.columns, start=start_col):
            column_letter = chr(64 + col_num)  # Convert column number to letter
            ws.column_dimensions[column_letter].width = 20  # Adjust as needed

        # Adjust row heights
        for row_num in range(start_row, start_row + len(df) + 1):  # Header + rows
            ws.row_dimensions[row_num].height = 18  # Adjust as needed

    def portfolio_risk_to_excel(
            self, 
            DIR: str,
            name: str,
            model: str,
            assets: List[str],
            weights: List[str],
            date_from: str,
            date_to: str,
        ) -> None:

        # Initialise portfolio risk data class and pull/calculate required data.
        portfolio_risk = PortfolioRiskData(model)
        portfolio_risk.get_data(assets, weights, date_from, date_to)

        factor_attribution = portfolio_risk.get_factor_attribution()
        factor_risk_proportion = portfolio_risk.get_factor_proportion_of_risk()
        factor_contribution_to_risk = (
            portfolio_risk.get_factor_contribution_to_risk(annualised=False)
        )
        stock_proportion_of_risk, stock_contribution_to_risk = (
            portfolio_risk.get_portfolio_risk_ts_by_stock(annualised=False)
        )
        risk_by_stock = portfolio_risk.get_factor_risk_by_stock(
            date_to, with_w = False, annualised=False
        )
        risk_by_stock_port,MCTR_by_stock_port,prop_by_stock_port = portfolio_risk.calculate_risk_port(date_to, annualised = False)

        factor_attribution_by_stock = (
            portfolio_risk.get_factor_attribution_by_stock_for_period(
                lookback=3 * 22
            )
        )
        exposures_by_stock = portfolio_risk.get_weighted_stock_exposures_for_date(
            date_to
        )

        final_portfolio_df = pd.DataFrame({'Assets': assets, 'Weights': weights, 'Direction': ['L' if w > 0 else 'S' for w in weights]})
        final_portfolio_df['Asset_direction'] = final_portfolio_df['Assets'] + '_' + final_portfolio_df['Direction']

        file_path = f'{DIR}/{name}_portfolio_{model}_{date_to}.xlsx'

        # Create a Pandas Excel writer using Openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write data to different sheets

            # Add the "Description" sheet
            self.add_description_sheet(writer.book)

            # Add the "weights" sheet
            final_portfolio_df.to_excel(writer, sheet_name='weights', index=False)

            portfolio_risk.exposures.to_excel(
                writer, sheet_name='exposures', index=True
            )
            factor_attribution.to_excel(
                writer, sheet_name='factor_attribution', index=True
            )
            portfolio_risk.risk_model_data.to_excel(
                writer, sheet_name='risk_and_return_ts', index=True
            )
            factor_risk_proportion.to_excel(
                writer, sheet_name='factor_risk_proportion', index=True
            )
            factor_contribution_to_risk.to_excel(
                writer, sheet_name='factor_risk_contribution', index=True
            )
            exposures_by_stock.to_excel(
                writer, sheet_name=f'exposures_{date_to}', index=True
            )
            factor_attribution_by_stock.to_excel(
                writer, sheet_name='factor_attribution_3m', index=True
            )
            risk_by_stock.to_excel(
                writer, sheet_name=f'single_stock_risk_{date_to}', index=True
            )
            risk_by_stock_port.to_excel(
                writer, sheet_name=f'port_stock_risk_{date_to}', index=True
            )
            MCTR_by_stock_port.to_excel(
                writer, sheet_name=f'port_stock_MCTR_{date_to}', index=True
            )
            prop_by_stock_port.to_excel(
                writer, sheet_name=f'port_stock_prop_{date_to}', index=True
            )
            stock_proportion_of_risk.to_excel(
                writer, sheet_name='stock_risk_proportion', index=True
            )
            stock_contribution_to_risk.to_excel(
                writer, sheet_name='stock_risk_contribution', index=True
            )

            factorset_df = pd.DataFrame(api_instance.get_risk_model_factorset(model))[['factor', 'descriptor']]
            self.add_factor_glossary_sheet(writer.book, factorset_df)

            # Add the "Disclaimer" sheet
            self.add_disclaimer_sheet(writer.book)



class PortfolioAnalysis():
    '''
    A class that takes the generated data from PortfolioRiskData
    to generate specific stats and charts required.
    '''
    def __init__(self, excel_file:str, factors:List=[]):
        self.factors = factors
        self.excel_file = excel_file
        self.latest_portfolio_exposure:Optional[pd.DataFrame] = None
        self.start_date:Optional[str] = None
        self.latest_date:Optional[str] = None
        self.three_month_portfolio_data:Optional[pd.DataFrame] = None
        self.three_month_portfolio_cumm_data:Optional[pd.DataFrame] = None
        self.three_month_portfolio_factor_attribution:Optional[
            pd.DataFrame
        ] = None
        self.three_month_portfolio_factor_attribution_cumm:Optional[
            pd.DataFrame
        ] = None
        self.three_months_factor_risk_contribution:Optional[
            pd.DataFrame
        ] = None
        self.latest_portfolio_factor_volatility:Optional[pd.DataFrame] = None
        self.max_abs_factor_exposure:Optional[str] = None
        self.second_max_abs_factor_exposure:Optional[str] = None
        self.filtered_latest_exposure_portfolio:Optional[pd.DataFrame] = None


    def _get_latest_portfolio_exposure(self):
        df = pd.read_excel(self.excel_file, sheet_name="exposures")
        df = df.set_index(df.columns[0]).iloc[-1]
        self.latest_date = str(df.name)
        df = df.to_frame('weight').sort_values(by="weight",ascending = False)
        if not self.latest_portfolio_exposure:
            self.latest_portfolio_exposure = df
        if not self.max_abs_factor_exposure:
            self.max_abs_factor_exposure = str(df['weight'].abs().idxmax())
        if not self.second_max_abs_factor_exposure:
            self.second_max_abs_factor_exposure = str(
                df['weight'].abs().drop(self.max_abs_factor_exposure).idxmax()
            )


    def _get_three_months_cumm_return_data(self):
        df = pd.read_excel(self.excel_file, sheet_name="risk_and_return_ts")
        df.set_index('date',inplace=True)
        df.index = pd.to_datetime(df.index)
        latest_date = df.index.max()
        three_months_back = latest_date - relativedelta(months=3)
        df = df[df.index >= three_months_back]
        df.index = df.index.strftime("%Y-%m-%d")
        self.start_date = df.index.min()

        if not self.three_month_portfolio_data:
            self.three_month_portfolio_data = df

        cumulative_df = (1 + df.iloc[:, :3]/100).cumprod() - 1
        df['total_risk'] = df['total_risk']*np.sqrt(252)

        df['factor_risk'] = (
            ((df['factor_risk']**2) / (df['total_risk'] / np.sqrt(252)) ** 2)
            * df['total_risk']
        )

        df['specific_risk'] = (
            ((df['specific_risk']**2) / (df['total_risk'] / np.sqrt(252)) ** 2)
            * df['total_risk']
        )

        result_df = pd.concat([cumulative_df, df.iloc[:, 3:]], axis=1)

        if not self.three_month_portfolio_cumm_data:
            self.three_month_portfolio_cumm_data = result_df

    def _get_three_months_cumm_factor_attribution(self):
        df = pd.read_excel(self.excel_file, sheet_name="factor_attribution")
        df.set_index(df.columns[0],inplace=True)
        df.index = pd.to_datetime(df.index)
        latest_date = df.index.max()
        three_months_back = latest_date - relativedelta(months=3)
        df = df[df.index >= three_months_back]
        df.index = df.index.strftime("%Y-%m-%d")
    
        if not self.three_month_portfolio_factor_attribution:
            self.three_month_portfolio_factor_attribution = df

        cumulative_df = (1 + df/100).cumprod() - 1

        if not self.three_month_portfolio_factor_attribution_cumm:
            self.three_month_portfolio_factor_attribution_cumm = cumulative_df

    def _get_three_months_factor_risk_contribution(self):
        df = pd.read_excel(
            self.excel_file,
            sheet_name="factor_risk_contribution"
        )
        df.set_index(df.columns[0],inplace=True)
        df.index = pd.to_datetime(df.index)
        latest_date = df.index.max()
        three_months_back = latest_date - relativedelta(months=3)
        df = df[df.index >= three_months_back]
        df.index = df.index.strftime("%Y-%m-%d")

        df = df * np.sqrt(252)

        df.drop(columns=['specific'],axis=1,inplace=True)
        if not self.three_months_factor_risk_contribution:
            self.three_months_factor_risk_contribution = df

    def _get_stock_exposure(self):
        if not isinstance(self.max_abs_factor_exposure, str):
            raise TypeError("max_abs_factor_exp should be a string")
        if not isinstance(self.second_max_abs_factor_exposure, str):
            raise TypeError("second_max_abs_factor_exp should be a string")
        
        df = pd.read_excel(
            self.excel_file,
            sheet_name= f"exposures_{self.latest_date}"
        )

        df.set_index(df.columns[0], inplace=True)

        if not self.factors:
            df = df[[
                self.max_abs_factor_exposure,
                self.second_max_abs_factor_exposure
            ]]
        else:
            if isinstance(self.latest_portfolio_exposure, pd.DataFrame):
                if (
                    isinstance(self.factors, List) 
                    and len(self.factors) == 2
                    and set(self.factors)
                    .issubset(set(list(self.latest_portfolio_exposure.index)))
                ):
                    df = df[self.factors]
                else:
                    raise TypeError('factors should be of a list of len 2')
        self.filtered_latest_exposure_portfolio = df

    def _get_latest_portfolio_factor_volatility(self):
        df = pd.read_excel(
            self.excel_file,
            sheet_name="stock_risk_contribution"
        )
        df = df.set_index(df.columns[0]).iloc[-1].to_frame().T
        df = df.loc[:, ~df.columns.str.contains('_specific')].T * np.sqrt(252)
        df = df.sort_values(by=df.columns[0],ascending = False)
        df.index = df.index.str.replace('_factor', '', regex=False)

        if not self.latest_portfolio_factor_volatility:
            self.latest_portfolio_factor_volatility = df

    def _create_bar_chart_for_latest_portfolio_factor_volatility(self):
        if not isinstance(
            self.latest_portfolio_factor_volatility,
            pd.DataFrame
        ):
            raise TypeError("filtered_portfolio should be a DataFrame")
        
        if isinstance(self.latest_date,str):
            latest_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
        else:
            raise TypeError("latest_date should be a string")
        
        date = latest_date.strftime("%d-%b-%y")
        df = self.latest_portfolio_factor_volatility
        if len(df) > 60:
            df = pd.concat([df.iloc[:30],df.iloc[-30:]])

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(
            df.index,
            df[df.columns[0]],
            color='steelblue',
              width=0.5
        )

        ax.set_title(
            (
                "                                                     "
                f"Portfolio Factor Vol attributable by stock ({date})"
                "                                                     "
            ),
            fontsize=13,
            fontweight='bold',
            style='italic',
            color='Black',
            pad=60,
            bbox=dict(
                facecolor='#CCE5FF',
                edgecolor='none',
                boxstyle='round,pad=0.8'
            )
        )

        ax.set_ylabel(
            "Share of % Portfolio Annualised Risk",
            fontsize=10,
            color='grey',
            fontweight='bold',
        )

        max_exposure = max(0, np.max(df.values))
        min_exposure = min(0, np.min(df.values))
        ax.set_ylim(min_exposure*1.3, max_exposure*1.3)

        ax.set_xticks(range(len(df.index)))
        ax.set_xticklabels(
            df.index,
            rotation=90, 
            fontsize=10,
            ha='center'
        )

        ax.tick_params(left=False, bottom=False)
        ax.axhline(0, color='grey', linewidth=0.5)

        for spine in ax.spines.values():
            spine.set_visible(False)

        plt.tight_layout()

        plt.savefig(f"portfolio_factor_vol.png")

    def _create_bar_chart_for_stock_factor_exposure(self):
        if not isinstance(
            self.filtered_latest_exposure_portfolio,
            pd.DataFrame
        ):
            raise TypeError("filtered_portfolio should be a DataFrame")
        
        if isinstance(self.latest_date,str):
            latest_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
        else:
            raise TypeError("latest_date should be a string")
        
        date = latest_date.strftime("%d-%b-%y")
        df = self.filtered_latest_exposure_portfolio

        df_map = pd.read_excel(self.excel_file, sheet_name="weights")
        name_map = dict(zip(df_map['Assets'], df_map['Asset_direction']))

        for col in df.columns:
            fig, ax = plt.subplots(figsize=(10, 6))
            df_ = df[col].to_frame(col).sort_values(
                by=col,
                ascending = False
            )
            df_.index = df_.index.map(name_map)
            if len(df_) > 60:
                df_ = pd.concat([df_.iloc[:30],df_.iloc[-30:]])
            bars = ax.bar(df_.index, df_[col], color='steelblue', width=0.5)

            ax.set_title(
                (
                    "                                                     "
                    f"Portfolio Exposures to {col} by stock as at {date}"
                    "                                                     "
                ),
                fontsize=13,
                fontweight='bold',
                style='italic',
                color='Black',
                pad=60,
                bbox=dict(
                    facecolor='#CCE5FF',
                    edgecolor='none',
                    boxstyle='round,pad=0.8'
                )
            )

            ax.set_ylabel(
                "Share of % Portfolio move for 1std move in factor",
                fontsize=10,
                color='grey',
                fontweight='bold',
            )

            max_exposure = max(0, np.max(df_.values))
            min_exposure = min(0, np.min(df_.values))
            ax.set_ylim(min_exposure*1.3, max_exposure*1.3)

            ax.set_xticks(range(len(df_.index)))
            ax.set_xticklabels(
                df_.index,
                rotation=90, 
                fontsize=10,
                ha='center'
            )

            ax.tick_params(left=False, bottom=False)
            ax.axhline(0, color='grey', linewidth=0.5)

            for spine in ax.spines.values():
                spine.set_visible(False)

            plt.tight_layout()
            
            plt.savefig(f"portfolio_exposures_to_{col}.png")

    def _create_bar_chart_for_latest_factor_attribution(self):
        if isinstance(
            self.three_month_portfolio_factor_attribution_cumm,
            pd.DataFrame
        ):

            df = self.three_month_portfolio_factor_attribution_cumm.iloc[-1]
            df.sort_values(ascending=False,inplace=True)

        else:
            raise TypeError(
                "three_month_portfolio_factor_attribution_cumm "
                "should be a dataframe"
            )
        if isinstance(self.latest_date,str):
            latest_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
        else:
            raise TypeError("latest_date should be a string")
        if isinstance(self.start_date,str):
            start_date = datetime.strptime(self.start_date, "%Y-%m-%d")
        else:
            raise TypeError("start_date should be a string")
        
        latest_date = latest_date.strftime("%d-%b-%y")
        start_date = start_date.strftime("%d-%b-%y")

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(df.index, list(df.values), color='steelblue', width=0.5)

        ax.set_title(
            (
                "                                                     "
                f"Factor Return Attribution ({start_date} to {latest_date})"
                "                                                     "
            ),
            fontsize=13,
            fontweight='bold',
            style='italic',
            color='Black',
            pad=60,
            bbox=dict(
                facecolor='#CCE5FF',
                edgecolor='none',
                boxstyle='round,pad=0.8'
            )
        )

        ax.set_ylabel(
            "% Return",
            fontsize=10,
            color='grey',
            fontweight='bold',
        )

        max_exposure = max(df.values)
        min_exposure = min(df.values)
        ax.set_ylim(min_exposure*1.6, max_exposure*1.6)

        ax.set_xticks(range(len(df.index)))
        ax.set_xticklabels(df.index, rotation=90, fontsize=10, ha='center')

        ax.tick_params(left=False, bottom=False)
        ax.axhline(0, color='grey', linewidth=0.5)

        ax.yaxis.set_major_formatter(PercentFormatter(xmax=1, decimals=2))

        for spine in ax.spines.values():
            spine.set_visible(False)

        plt.tight_layout()

        plt.savefig("portfolio_factor_attribution.png")

    def _create_line_chart_for_three_months_cumm_return(self):
        if isinstance(self.three_month_portfolio_cumm_data, pd.DataFrame):

            total_return = (
                self.three_month_portfolio_cumm_data
                ['total_return'].to_list()
            )
            factor_return = (
                self.three_month_portfolio_cumm_data
                ['factor_return'].to_list()
            )
            specific_return = (
                self.three_month_portfolio_cumm_data
                ['specific_return'].to_list()
            )

            dates = (
                self.three_month_portfolio_cumm_data.index.to_list()
            )

            formatted_dates = pd.to_datetime(dates).strftime("%d-%b")
        else:
            raise TypeError(
                "three_month_portfolio_cumm_data should be a dataframe"
            )
        
        fig, ax = plt.subplots(figsize=(10, 6))
        ax2 = ax.twinx()
        ar, = ax.plot(
            formatted_dates,
            total_return,
            label='Actual',
            linewidth=1.5,
            color="blue",
        )
        fr, = ax2.plot(
            formatted_dates,
            factor_return,
            label='Factor (rhs)',
            linewidth=1.5,
            color="#8B0000",
        )
        sr, = ax.plot(
            formatted_dates,
            specific_return,
            label='Specific',
            linewidth=1.5,
            color="#32CD32",
        )
        ax.set_title(
            (
                "                                                          "
                f"Cumulative Return Attribution"
                "                                                          "
            ),
            fontsize=13,
            fontweight='bold',
            style='italic',
            color='Black',
            pad=60,
            bbox=dict(
                facecolor='#CCE5FF',
                edgecolor='none',
                boxstyle='round,pad=0.8'
            )
        )
        
        lines = [ar, fr, sr]
        labels = [line.get_label() for line in lines]
        ax.legend(
            lines,
            labels,
            loc='upper center',
            bbox_to_anchor=(0.5, 1.1),
            ncol=3,
            frameon=False,
            fontsize=10
        )

        num_ticks = 17
        tick_positions = np.linspace(
            0, len(formatted_dates) - 1, num_ticks, dtype=int
        )
        ax.set_xticks(tick_positions)
        ax.set_xticklabels(
            [formatted_dates[i] for i in tick_positions], rotation=90
        )

        ax.tick_params(left=False, bottom=False)
        ax2.tick_params(left=False, bottom=False)
        ax2.tick_params(axis='y', which='both', length=0)
        ax.axhline(0, color='grey', linewidth=0.5)

        max_exposure = max(
            max(total_return),
            max(specific_return)
        )
        min_exposure = min(
            min(total_return),
            min(specific_return)
        )
        if abs(max_exposure) > abs(min_exposure):
            ax.set_ylim(-(abs(max_exposure)*1.3), abs(max_exposure)*1.3)
        else:
            ax.set_ylim(-(abs(min_exposure)*1.3), abs(min_exposure)*1.3)

        max_factor = max(factor_return)
        min_factor = min(factor_return)

        if abs(max_factor) > abs(min_factor):
            ax2.set_ylim(-(abs(max_factor)*1.3), abs(max_factor)*1.3)
        else:
            ax2.set_ylim(-(abs(min_factor)*1.3), abs(min_factor)*1.3)

        ax.yaxis.set_major_formatter(PercentFormatter(xmax=1, decimals=2))
        ax2.yaxis.set_major_formatter(PercentFormatter(xmax=1, decimals=2))

        ax.set_xlim(left=0)

        for spine in ax.spines.values():
            spine.set_visible(False)

        for spine in ax2.spines.values():
            spine.set_visible(False)

        plt.tight_layout()

        plt.savefig("LineChart.png")

    def _create_area_chart_for_three_months_risk(self):
        if isinstance(self.three_month_portfolio_cumm_data, pd.DataFrame):

            factor_risk = (
                self.three_month_portfolio_cumm_data
                ['factor_risk'].to_list()
            )
            specific_risk = (
                self.three_month_portfolio_cumm_data
                ['specific_risk'].to_list()
            )

            dates = (
                self.three_month_portfolio_cumm_data.index.to_list()
            )

            formatted_dates = pd.to_datetime(dates).strftime("%d-%b")
        else:
            raise TypeError(
                "three_month_portfolio_cumm_data should be a dataframe"
            )
        if isinstance(self.latest_date,str):
            latest_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
        else:
            raise TypeError("latest_date should be a string")
        if isinstance(self.start_date,str):
            start_date = datetime.strptime(self.start_date, "%Y-%m-%d")
        else:
            raise TypeError("start_date should be a string")
        
        latest_date = latest_date.strftime("%d-%b-%y")
        start_date = start_date.strftime("%d-%b-%y")

        fig, ax = plt.subplots(figsize=(10, 6))

        ax.stackplot(
            formatted_dates,
            factor_risk,
            specific_risk,
            labels=['factor_risk', 'specific_risk'],
            colors=['#8B0000', '#32CD32'],
            alpha=0.8
        )

        ax.set_title(
                (
                    "                                                     "
                    "Predicted Annualised Risk Attribution"
                    f"({start_date} to {latest_date})"
                    "                                                     "
                ),
                fontsize=13,
                fontweight='bold',
                style='italic',
                color='Black',
                pad=60,
                bbox=dict(
                    facecolor='#CCE5FF',
                    edgecolor='none',
                    boxstyle='round,pad=0.8'
                )
            )

        ax.set_ylabel(
            "% Annualised Risk",
            fontsize=10,
            color='grey',
            fontweight='bold',
        )
        max_exposure = max(
            max(factor_risk),
            max(specific_risk)
        )
        min_exposure = min(
            0,
            min(factor_risk),
            min(specific_risk)
        )
        max_exposure = max_exposure * 1.6
        min_exposure = min_exposure * 1.3
        ax.set_ylim(min_exposure, max_exposure)

        y_ticks = np.linspace(min_exposure, max_exposure, 9)
        ax.set_yticks(y_ticks)
        ax.set_yticklabels([f"{tick:.1f}" for tick in y_ticks])

        ax.tick_params(left=False, bottom=False)
        ax.axhline(0, color='grey', linewidth=0.5)

        num_ticks = 17
        tick_positions = np.linspace(
            0, len(formatted_dates) - 1, num_ticks, dtype=int
        )
        ax.set_xticks(tick_positions)
        ax.set_xticklabels(
            [formatted_dates[i] for i in tick_positions], rotation=90
        )

        ax.legend(
            loc='upper center',
            bbox_to_anchor=(0.5, 1.1),
            ncol=2,
            frameon=False
        )

        for spine in ax.spines.values():
            spine.set_visible(False)

        plt.tight_layout()
        plt.savefig("predicted_annualised_risk.png")

    def _create_area_chart_for_three_months_factor_risk(self):
            if isinstance(
                self.three_months_factor_risk_contribution,
                pd.DataFrame
            ):
                df = self.three_months_factor_risk_contribution

                dates = (
                    df.index.to_list()
                )

                formatted_dates = pd.to_datetime(dates).strftime("%d-%b")
            else:
                raise TypeError(
                    "three_months_factor_risk_contribution must be a dataframe"
                )
            if isinstance(self.latest_date,str):
                latest_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
            else:
                raise TypeError("latest_date should be a string")
            if isinstance(self.start_date,str):
                start_date = datetime.strptime(self.start_date, "%Y-%m-%d")
            else:
                raise TypeError("start_date should be a string")
            
            latest_date = latest_date.strftime("%d-%b-%y")
            start_date = start_date.strftime("%d-%b-%y")

            try:
                colors = [COLORS[col] for col in df.columns]
            except KeyError as e:
                raise KeyError(f"Missing color mapping for column: {e}")

            fig, ax = plt.subplots(figsize=(10, 6))

            df_positive = df[df >= 0].fillna(0).T
            df_negative = df[df < 0].fillna(0).T
            ax.stackplot(df.index, df_positive, colors=colors, alpha=0.8)
            ax.stackplot(df.index, df_negative, colors=colors, alpha=0.8) 

            ax.set_title(
                    (
                        "                                                     "
                        "          Factor Risk Attribution"
                        f"({start_date} to {latest_date})          "
                        "                                                     "
                    ),
                    fontsize=13,
                    fontweight='bold',
                    style='italic',
                    color='Black',
                    pad=60,
                    bbox=dict(
                        facecolor='#CCE5FF',
                        edgecolor='none',
                        boxstyle='round,pad=0.8'
                    )
                )

            ax.set_ylabel(
                "% Annualised Risk",
                fontsize=10,
                color='grey',
                fontweight='bold',
            )
            max_exposure = max(
                0,
                df_positive.sum(axis=0).max()
            )
            min_exposure = min(
                0,
                df_negative.sum(axis=0).min()
            )
            max_exposure = max_exposure * 1.1
            min_exposure = min_exposure * 1.1
            ax.set_ylim(min_exposure, max_exposure)

            y_ticks = np.linspace(min_exposure, max_exposure, 9)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels([f"{tick:.1f}" for tick in y_ticks])

            ax.tick_params(left=False, bottom=False)

            num_ticks = 17
            tick_positions = np.linspace(
                0, len(formatted_dates) - 1, num_ticks, dtype=int
            )
            ax.set_xticks(tick_positions)
            ax.set_xticklabels(
                [formatted_dates[i] for i in tick_positions], rotation=90
            )

            ax.legend(
                labels = df.columns,
                loc='upper center',
                bbox_to_anchor=(0.5, -0.2),
                ncol=5,
                frameon=False,
                fontsize=8
            )

            for spine in ax.spines.values():
                spine.set_visible(False)

            plt.tight_layout()
            plt.savefig("factor_risk_attribution.png")

    def _create_bar_chart_for_latest_portfolio_exposure(self):
        if isinstance(self.latest_portfolio_exposure, pd.DataFrame):

            factors = self.latest_portfolio_exposure.index.to_list()
            exposures = self.latest_portfolio_exposure['weight'].to_list()
        
        else:
            raise TypeError("latest_portfolio_exposure should be a dataframe")

        if isinstance(self.latest_date,str):
            formated_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
        else:
            raise TypeError("latest_date should be a string")
        
        formated_date = formated_date.strftime("%d-%b-%y")

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(factors, exposures, color='steelblue', width=0.5)

        ax.set_title(
            (
                "                                                     "
                f"Portfolio Factor Exposures as at {formated_date}"
                "                                                     "
            ),
            fontsize=13,
            fontweight='bold',
            style='italic',
            color='Black',
            pad=60,
            bbox=dict(
                facecolor='#CCE5FF',
                edgecolor='none',
                boxstyle='round,pad=0.8'
            )
        )

        ax.set_ylabel(
            "% Portfolio Move for 1 std daily\n"
                "move in Factor",
            fontsize=10,
            color='grey',
            fontweight='bold',
        )

        max_exposure = max(exposures)
        min_exposure = min(exposures)
        ax.set_ylim(min_exposure*1.3, max_exposure*1.3)

        ax.set_xticks(range(len(factors)))
        ax.set_xticklabels(factors, rotation=90, fontsize=10, ha='center')

        ax.tick_params(left=False, bottom=False)
        ax.axhline(0, color='grey', linewidth=0.5)

        for spine in ax.spines.values():
            spine.set_visible(False)

        plt.tight_layout()

        plt.savefig("portfolio_exposures.png")

    def _create_plot_chart(self):
        if isinstance(self.three_month_portfolio_data, pd.DataFrame):
            total_return = (
                self.three_month_portfolio_data
                ['total_return'].to_list()
            )
            factor_return = (
                self.three_month_portfolio_data
                ['factor_return'].to_list()
            )
        else:
            raise TypeError("three_m_portfolio_data should be a DataFrame")
        
        if isinstance(self.latest_date,str):
            latest_date = datetime.strptime(self.latest_date, "%Y-%m-%d")
        else:
            raise TypeError("latest_date should be a string")
        if isinstance(self.start_date,str):
            start_date = datetime.strptime(self.start_date, "%Y-%m-%d")
        else:
            raise TypeError("start_date should be a string")
        
        latest_date = latest_date.strftime("%d-%b-%y")
        start_date = start_date.strftime("%d-%b-%y")
        
        fig, ax = plt.subplots(figsize=(10, 6))

        ax.scatter(
            total_return,
            factor_return,
            color="steelblue",
            alpha=0.7
        )

        max_abs_x = max(abs(min(total_return)), abs(max(total_return))) 
        max_abs_y = max(abs(min(factor_return)), abs(max(factor_return)))

        ax.set_xlim(-max_abs_x*1.3, max_abs_x*1.3)
        ax.set_ylim(-max_abs_y*1.3, max_abs_y*1.3)

        ax.set_title(
            (
                "                                                     "
                "Actual Daily Returns vs. Factor Daily Returns "
                f"({start_date} to {latest_date})"
                "                                                     "
            ),
            fontsize=13,
            fontweight='bold',
            style='italic',
            color='Black',
            pad=60,
            bbox=dict(
                facecolor='#CCE5FF',
                edgecolor='none',
                boxstyle='round,pad=0.8'
            )
        )

        ax.set_ylabel(
            "Factor Daily % Return",
            fontsize=10,
            color='grey',
            fontweight='bold',
        )
        ax.yaxis.set_label_coords(-0.05, 0.5)

        ax.set_xlabel(
            "Actual Daily % Return",
            fontsize=10,
            color='grey',
            fontweight='bold',
        )
        ax.xaxis.set_label_coords(0.5, -0.05)

        ax.yaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=2))
        ax.xaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=2))

        coefficients = np.polyfit(total_return, factor_return, 1)
        trendline = np.poly1d(coefficients)
        x_range = np.linspace(min(total_return), max(total_return), 100)
        ax.plot(
            x_range,
            trendline(x_range),
            color="steelblue",
            linestyle="dotted"
        )

        ax.spines['left'].set_position(('data', 0))
        ax.spines['bottom'].set_position(('data', 0))

        ax.tick_params(left=False, bottom=False)

        for spine in ax.spines.values():
            spine.set_visible(False)

        ax.grid(alpha=0.3)

        factor_return = np.array(factor_return)
        y_pred = trendline(total_return)
        ss_total = np.sum((factor_return - np.mean(factor_return))**2)
        ss_residual = np.sum((factor_return - y_pred)**2)
        r_squared = 1 - (ss_residual / ss_total)
        if coefficients[1] < 0:
            sign = "-"
        else:
            sign = "+"
        equation_text = (
            f"$y = {coefficients[0]:.4f}x {sign} "
            f"{abs(coefficients[1]):.4f}$\n$R^2 = {r_squared:.4f}$"
        )
        ax.text(
            0.95,
            0.95,
            equation_text,
            transform=ax.transAxes,
            fontsize=8,
            verticalalignment='top'
        )

        plt.tight_layout()
        plt.savefig("scatter_plot.png")

    def _export_to_excel(self):
        wb = openpyxl.load_workbook(self.excel_file)
        sheet_name = "charts"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        charts_sheet = wb[sheet_name]
        wb.move_sheet(charts_sheet, offset=-len(wb.sheetnames) + 2)
        if not self.factors:
            chart_files = [
                "LineChart.png",
                "portfolio_exposures.png",
                f"portfolio_exposures_to_{self.max_abs_factor_exposure}.png",
                "scatter_plot.png",
                "predicted_annualised_risk.png",
                f"portfolio_exposures_to_{self.second_max_abs_factor_exposure}.png",
                "portfolio_factor_attribution.png",
                "factor_risk_attribution.png",
                "portfolio_factor_vol.png"
            ]
        else:
            chart_files = [
                "LineChart.png",
                "portfolio_exposures.png",
                f"portfolio_exposures_to_{self.factors[0]}.png",
                "scatter_plot.png",
                "predicted_annualised_risk.png",
                f"portfolio_exposures_to_{self.factors[1]}.png",
                "portfolio_factor_attribution.png",
                "factor_risk_attribution.png",
                "portfolio_factor_vol.png"
            ]
        start_row, start_col = 3, 2
        row_offset, col_offset = 32, 16
        for i, chart_file in enumerate(chart_files):
            img = Image(chart_file)
            row = start_row + (i // 3) * row_offset
            col = start_col + (i % 3) * col_offset
            col_letter = get_column_letter(col)
            cell = f"{col_letter}{row}"
            ws.add_image(img, cell)

        ws.sheet_view.showGridLines = False

        wb.save(self.excel_file)

    def runner(self):
        self._get_latest_portfolio_exposure()
        self._create_bar_chart_for_latest_portfolio_exposure()
        self._get_three_months_cumm_return_data()
        self._create_line_chart_for_three_months_cumm_return()
        self._create_area_chart_for_three_months_risk()
        self._create_plot_chart()
        self._get_three_months_cumm_factor_attribution()
        self._create_bar_chart_for_latest_factor_attribution()
        self._get_three_months_factor_risk_contribution()
        self._create_area_chart_for_three_months_factor_risk()
        self._get_stock_exposure()
        self._create_bar_chart_for_stock_factor_exposure()
        self._get_latest_portfolio_factor_volatility()
        self._create_bar_chart_for_latest_portfolio_factor_volatility()
        self._export_to_excel()


class RiskData(ABC):
    """
    Abstract base class for retrieving risk data for an asset via Qi's API, and has
    several methods to derive further risk metrics (return attribtuions and vol/risk predictions).

    Attributes:
        model (str): The risk model used to retrieve risk data.
        exposures (pd.DataFrame): Asset exposures to various risk factors.
        exposure_errors (pd.DataFrame): Errors associated with asset exposures.
        risk_model_data (pd.DataFrame): Risk model data for the asset.
        factor_returns (pd.DataFrame): Factor returns data for the factors in the risk model.
        factor_stds (pd.DataFrame): Rolling standard deviations for risk factors.
        factor_risk (pd.DataFrame): Factor risk values for the asset.
        factor_covariance (Dict[str, Dict[str, Dict[str, float]]]): Covariance matrix for factors.
    """

    def __init__(self, model: str):
        self.model = model
        self.api_data = ApiData()
        self.exposures: pd.DataFrame = pd.DataFrame()
        self.exposure_errors: pd.DataFrame = pd.DataFrame()
        self.risk_model_data: pd.DataFrame = pd.DataFrame()
        self.factor_returns: pd.DataFrame = pd.DataFrame()
        self.factor_stds: pd.DataFrame = pd.DataFrame()
        self.factor_risk: pd.DataFrame = pd.DataFrame()
        self.factor_covariance: Dict[str, Dict[str, Dict[str, float]]] = {}

    @abstractmethod
    def get_data(self, **kwags) -> None:
        """
        Abstract method to retrieve data for the asset.

        This method must be implemented by subclasses to define the specific data retrieval
        logic based on the asset type, model, and other parameters.

        Args:
            **kwargs: Arbitrary keyword arguments for data retrieval, which may vary based on the subclass.

        Raises:
            NotImplementedError: If not implemented by a subclass.
        """
        raise NotImplementedError()

    @abstractmethod
    def calculate_risk(self) -> pd.DataFrame:
        """
        Abstract method to calculate the risk for an asset.

        Subclasses must implement this method to provide calculations for risk forecasts
        based on the asset and model.

        Returns:
            pd.DataFrame: DataFrame containing factor risk values.

        Raises:
            NotImplementedError: If not implemented by a subclass.
        """
        raise NotImplementedError()

    def get_factor_attribution(self) -> pd.DataFrame:
        """
        Calculates factor attributions, which represent the contributions of each factor to the asset's returns.
        For factor return attributions we use exposures on t-1, and factor returns on t, to calculate factor
        attributions on t.

        Returns:
            pd.DataFrame: A DataFrame containing factor attributions.
        """
        factor_attribution = self.exposures.shift(1) * self.factor_returns
        return factor_attribution

    @staticmethod
    def _get_proportion_of_risk(
        risk: [pd.DataFrame, pd.Series], total_risk: pd.Series
    ) -> pd.DataFrame:
        """
        Helper function to calculate the proportion of risk attributable to each factor (and specific).

        Args:
            risk [pd.DataFrame, pd.Series]: Either a DataFrame or Series of risk values.
            total_risk: A data series of total risk values.

        Returns:
            pd.DataFrame: A DataFrame containing the proportion of risk by factor.
        """
        risk = pd.DataFrame(risk)
        risk_proportion = (
            (risk**2 * np.sign(risk)).div(total_risk**2, axis=0).dropna()
        )

        return risk_proportion

    def get_factor_proportion_of_risk(self) -> pd.DataFrame:
        """
        Calculates the proportion of risk attributable to each factor (and specific).

        Returns:
            pd.DataFrame: A DataFrame containing the proportion of risk by factor.
        """
        factor_risk_proportion = self._get_proportion_of_risk(
            self.factor_risk, self.risk_model_data.total_risk
        )

        specific_risk_proportion = self._get_proportion_of_risk(
            self.risk_model_data.specific_risk, self.risk_model_data.total_risk
        )
        factor_risk_proportion['specific'] = specific_risk_proportion[
            'specific_risk'
        ]

        return factor_risk_proportion

    @staticmethod
    def _get_contribution_to_risk(
        risk_proportion: pd.DataFrame,
        total_risk: pd.Series,
        annualised: bool = False,
    ) -> pd.DataFrame:
        """
        Helper function to calculate the contribution of each factor (and specific) to the overall risk of the asset.

        Args:
            risk_proportion (pd.DataFrame): DataFrame of factor risk proportions.
            total_risk (pd.Series): data Series of total risk values.
            annualised (bool, optional): Whether to annualize the risk contributions. Default is False.

        Returns:
            pd.DataFrame: A DataFrame containing the factor contributions to risk.
        """
        scalar = 252**0.5 if annualised else 1
        contribution_to_risk = risk_proportion.mul(
            total_risk * scalar, axis=0
        ).dropna()

        return contribution_to_risk

    def get_factor_contribution_to_risk(
        self, annualised: bool = False
    ) -> pd.DataFrame:
        """
        Calculates the contribution of each factor (and specific) to the overall risk of the asset.

        Args:
            annualised (bool, optional): Whether to annualize the risk contributions. Default is False.

        Returns:
            pd.DataFrame: A DataFrame containing the factor contributions to risk.
        """
        factor_risk_proportion = self.get_factor_proportion_of_risk()
        factor_contribution_to_risk = self._get_contribution_to_risk(
            factor_risk_proportion, self.risk_model_data.total_risk, annualised
        )

        return factor_contribution_to_risk


class AssetRiskData(RiskData):

    def get_data(self, asset: str, date_from: str, date_to: str):
        """
        Retrieves risk data for a given asset over the specified time range.

        Args:
            asset (str): The asset for which risk data is being retrieved.
            date_from (str): The start date for the data in 'YYYY-MM-DD' format.
            date_to (str): The end date for the data in 'YYYY-MM-DD' format.
        """
        self.exposures = self.api_data.get_exposure_data(
            self.model, asset, date_from, date_to
        )
        self.exposure_errors = self.api_data.get_exposure_error_data(
            self.model, asset, date_from, date_to
        )
        self.risk_model_data = self.api_data.get_risk_model_data(
            self.model, asset, date_from, date_to
        )
        self.factor_returns = self.api_data.get_factor_returns_data(
            self.model, date_from, date_to
        )
        self.factor_stds = self.api_data.get_factor_stds_data(
            self.model, date_from, date_to
        )
        self.factor_covariance = self.api_data.get_factor_covariance_data(
            self.model, date_from, date_to
        )
        print(f'Fetched data for {asset}, in model {self.model}')

        self.factor_risk = self.calculate_risk()

    def calculate_risk(self, annualised: bool = False) -> pd.DataFrame:
        """
        Calculates the factor risk for each date, optionally annualized.

        Total_risk = (factor_risk**2 + specific_risk**2)**0.5
        factor_risk**2 = X.F.X  ,
        where X is exposures, F is covariance matrix

        Args:
            annualised (bool, optional): Whether to annualize the risk values. Default is False.

        Returns:
            pd.DataFrame: A DataFrame containing factor risk values for each date.
        """
        factor_results = []

        for date in self.exposures.index:
            date_exposures = self.exposures.loc[date]
            date_covar_matrix = pd.DataFrame(self.factor_covariance[date])
            date_exposures = date_exposures[date_covar_matrix.index]
            factor_risks = (
                date_exposures.dot(date_covar_matrix)
            ) * date_exposures

            scalar = 252**0.5 if annualised else 1

            factor_result = {'date': date}
            for factor, value in factor_risks.items():
                if value < 0:
                    value = -1 * (abs(value) ** 0.5) * scalar
                else:
                    value = (value**0.5) * scalar
                factor_result.update({factor: value})
            factor_results.append(factor_result)

        factor_risk_df = pd.DataFrame(factor_results).set_index('date')

        return factor_risk_df


class ApiData:
    """
    Class that provides methods to retrieve data using Qi's API.
    """

    @staticmethod
    def _get_data_for_func(
        func: Callable,
        model: str,
        start: str,
        end: str,
        asset: str = None,
    ) -> Dict:
        """
        Helper method to retrieve data for a given function and time period, splitting the data
        by year to handle large date ranges.

        Args:
            func (Callable): The API function to be called for data retrieval.
            model (str): The risk model used.
            start (str): The start date in 'YYYY-MM-DD' format.
            end (str): The end date in 'YYYY-MM-DD' format.
            asset (str, optional): The asset for which data is being retrieved. Default is None.

        Returns:
            Dict: The retrieved data as a dictionary.
        """
        year_start = int(start[:4])
        year_end = int(end[:4])
        data = {}

        for year in range(year_start, year_end + 1):

            if year != year_start:
                date_from = '%d-01-01' % year
            else:
                date_from = start

            if year != year_end:
                date_to = '%d-12-31' % year
            else:
                date_to = end

            if asset:
                data.update(
                    func(
                        model,
                        instrument=asset,
                        date_from=date_from,
                        date_to=date_to,
                    )
                )

            else:
                data.update(
                    func(
                        model,
                        date_from=date_from,
                        date_to=date_to,
                    )
                )

        return data

    def get_exposure_data(
        self, model: str, asset: str, date_from: str, date_to: str
    ) -> pd.DataFrame:
        func = api_instance.get_exposures_for_risk_model
        exposures = self._get_data_for_func(
            func, model, date_from, date_to, asset
        )

        return pd.DataFrame.from_dict(exposures).T

    def get_exposure_error_data(
        self, model: str, asset: str, date_from: str, date_to: str
    ) -> pd.DataFrame:
        func = api_instance.get_exposure_errors_for_risk_model
        exposure_errors = self._get_data_for_func(
            func, model, date_from, date_to, asset
        )

        return pd.DataFrame.from_dict(exposure_errors).T

    def get_risk_model_data(
        self, model: str, asset: str, date_from: str, date_to: str
    ) -> pd.DataFrame:
        func = api_instance.get_risk_model_data_for_risk_model
        risk_model_data = self._get_data_for_func(
            func, model, date_from, date_to, asset
        )

        return pd.DataFrame.from_dict(risk_model_data).T

    def get_factor_returns_data(
        self, model: str, date_from: str, date_to: str
    ) -> pd.DataFrame:
        func = api_instance.get_factor_returns_for_risk_model
        factor_returns = self._get_data_for_func(
            func, model, date_from, date_to
        )

        return pd.DataFrame.from_dict(factor_returns).T

    def get_factor_stds_data(
        self, model: str, date_from: str, date_to: str
    ) -> pd.DataFrame:
        func = api_instance.get_factor_returns_stds_for_risk_model
        factor_stds = self._get_data_for_func(func, model, date_from, date_to)

        return pd.DataFrame.from_dict(factor_stds).T

    def get_factor_covariance_data(
        self, model: str, date_from: str, date_to: str
    ) -> Dict[str, Dict[str, Dict[str, float]]]:
        func = api_instance.get_covariances_for_risk_model
        cov_data = self._get_data_for_func(func, model, date_from, date_to)

        return cov_data

    def get_universe_by_model(
        self, risk_model, identifier_type, include_delisted
    ):
        api_response = pd.DataFrame(api_instance.get_risk_model_universe(
            risk_model,
            identifier_type=identifier_type,
            include_delisted=include_delisted,
        )).T[identifier_type].tolist()

        return api_response

class PortfolioRiskData(RiskData):
    """
    A subclass of RiskData that calculates and manages portfolio risk data for a
    portfolio of assets. This class retrieves risk data for a portfolio, calculates
    exposures, risk, and returns on a portfolio level, and provides methods to analyze
    risk contributions by factor and asset.
    """

    def get_data(
        self,
        assets: List[str],
        weights: List[float],
        date_from: str,
        date_to: str,
    ) -> None:
        """
        Retrieves risk data for a given list of assets and weights
        in a portfolio over the specified time range.

        Args:
            assets List[str]: list of assets in portfolio.
            weights List[float]: list of portfolio weights.
            date_from (str): The start date for the data in 'YYYY-MM-DD' format.
            date_to (str): The end date for the data in 'YYYY-MM-DD' format.

        Raises:
            AssertionError: If the number of assets does not match the number of weights.

        Returns:
            None: The method retrieves and stores risk data internally for further calculations.
        """
        assert len(weights) == len(
            assets
        ), 'array of assets must be of same length as weights'
        self.weights = weights
        self.asset_data: Dict[str, AssetRiskData] = {}

        # Assuming 'assets' is your list of assets and 'self.model', 'date_from', 'date_to' are defined
        results = Parallel(n_jobs=10)(
            delayed(self.process_asset_data)(
                asset, self.model, date_from, date_to
            )
            for asset in assets
        )

        # Collect results into the dictionary
        self.asset_data = {asset: data for asset, data in results}

        self.exposures = self.calculate_portfolio_exposures()

        self.factor_returns = self.api_data.get_factor_returns_data(
            self.model, date_from, date_to
        )
        self.factor_stds = self.api_data.get_factor_stds_data(
            self.model, date_from, date_to
        )
        self.factor_covariance = self.api_data.get_factor_covariance_data(
            self.model, date_from, date_to
        )
        self.exposures = self.exposures.dropna()
        self.factor_risk = self.calculate_risk()

    def process_asset_data(self, asset, model, date_from, date_to):
        asset_risk_data = AssetRiskData(model)
        asset_risk_data.get_data(asset, date_from, date_to)
        return asset, asset_risk_data

    def calculate_portfolio_exposures(self) -> pd.DataFrame:
        """
        Calculates weighted exposures for each factor based on
        individual asset exposures and portfolio weights.

        Returns:
            pd.DataFrame: A DataFrame containing the weighted portfolio exposures for each factor.
        """

        weighted_exposures = sum(
            [
                x.exposures * self.weights[i]
                for i, x in enumerate(self.asset_data.values())
            ]
        )

        return weighted_exposures

    def calculate_risk_port(self,date: str, annualised: bool = False) -> pd.DataFrame:		 
        """		 
        Calculates the breakdown per stock of the factor risk for each date, optionally annualized.		 
        For further information please see the explanatory spreadsheet "Volat Corr" breakdown per stock		 
                
        Returns:		 
            pd.DataFrame: DataFrame that contains the factor risk values for each stock.		 
        """		 
        weights_df = pd.DataFrame(self.weights, columns=['weights'],index=self.asset_data.keys())		 
        port_date_exposures = self.exposures.loc[date]		 
        date_covar_matrix = pd.DataFrame(self.factor_covariance[date])		 
        temp_risk_vect = port_date_exposures.dot(date_covar_matrix)		 
        stock_risk_dic = {}		 
        stock_risk2_dic = {}		 
        i=0		 
        specific_risk_t= np.empty(len(weights_df))		 
        for stock,s_weight in weights_df.iterrows():		 
            w = s_weight['weights'] * (252.0 if annualised else 1.0)		 
            t2 = (self.asset_data[stock].exposures.loc[date] * temp_risk_vect) * w		 
            stock_risk2_dic[stock] = t2		 
            stock_risk_dic[stock] = np.sign(t2) * np.sqrt(np.abs(t2))		 
            specific_risk_t[i] = self.asset_data[stock].risk_model_data.specific_risk[date] * w		 
            i=i+1		 
        stock_risk_df = pd.DataFrame.from_dict(stock_risk_dic,orient='index',columns=self.exposures.columns.values.tolist())		 
        stock_risk_prop_df = pd.DataFrame.from_dict(stock_risk2_dic,orient='index',columns=self.exposures.columns.values.tolist())		 
        stock_risk_df = stock_risk_df.assign(specific_risk = specific_risk_t)		 
        stock_risk_prop_df = stock_risk_prop_df.assign(specific_risk = specific_risk_t**2)		 
        total_risk2 = float(stock_risk_prop_df.sum(axis=0).sum())		 
                
        return stock_risk_df,stock_risk_prop_df.div(np.sqrt(total_risk2)),stock_risk_prop_df.div(total_risk2)

    def calculate_risk(self) -> pd.DataFrame:
        """
        Calculates the factor risk and risk_model_data (total, specific and factor risk and return)
        for each date, optionally annualized (just applies for risk values).

        Total_risk = (factor_risk**2 + specific_risk**2)**0.5
        factor_risk**2 = X . F . X  ,
        where X is weighted portfolio exposures, F is covariance matrix
        specific_risk**2 = w.T . R . w
        where w is vector of portfolio weights, R is diagonal matrox of specific risk variances.

        Returns:
            pd.DataFrame: DataFrame that contains the factor risk values for each date.
        """

        weights_df = pd.DataFrame(self.weights, index=self.asset_data.keys())
        portfolio_factor_returns = self.get_factor_attribution().T.sum()

        factor_results = []
        aggregate_results = []
        for date in self.exposures.index:
            date_exposures = self.exposures.loc[date]
            date_covar_matrix = pd.DataFrame(self.factor_covariance[date])
            factor_risks = (
                self.exposures.loc[date].dot(date_covar_matrix)
            ) * self.exposures.loc[date]

            diag_specific_risk_matrix = np.diag(
                [
                    x.risk_model_data.specific_risk[date] ** 2
                    for x in self.asset_data.values()
                ]
            )
            weighted_diag_specific_risk_matrix = np.dot(
                np.dot(weights_df.T, diag_specific_risk_matrix), weights_df
            )

            factor_var = factor_risks.sum()
            specific_var = weighted_diag_specific_risk_matrix[ZERO_IDX][
                ZERO_IDX
            ]

            factor_risk = factor_var**0.5
            specific_risk = specific_var**0.5
            total_risk = (factor_var + specific_var) ** 0.5

            total_return = sum(
                [
                    x.risk_model_data.total_return[date] * self.weights[i]
                    for i, x in enumerate(self.asset_data.values())
                ]
            )
            factor_return = portfolio_factor_returns[date]
            specific_return = total_return - factor_return

            aggregate_results.append(
                {
                    'date': date,
                    'total_return': total_return,
                    'factor_return': factor_return,
                    'specific_return': specific_return,
                    'total_risk': total_risk,
                    'factor_risk': factor_risk,
                    'specific_risk': specific_risk,
                }
            )

            factor_result = {'date': date}
            for factor, value in factor_risks.items():
                if value < 0:
                    value = -1 * (abs(value) ** 0.5)
                else:
                    value = value**0.5
                factor_result.update({factor: value})
            factor_results.append(factor_result)

        factor_risk_df = pd.DataFrame(factor_results).set_index('date')
        risk_df = pd.DataFrame(aggregate_results).set_index('date')

        self.risk_model_data = risk_df

        return factor_risk_df

    def get_portfolio_risk_ts_by_stock(
        self, annualised: bool = False
    ) -> [pd.DataFrame, pd.DataFrame]:
        """
        Calculates the time series of portfolio risk due to each stock in the portfolio,
        split into factor and specific risks from each stock.

        Args:
            annualised (bool, optional): Whether to annualize the risk values. Defaults to False.

        Returns:
            Tuple[pd.DataFrame, pd.DataFrame]:
                - The first DataFrame contains the proportion of portfolio risk attributable to each stock.
                - The second DataFrame contains each stock's contribution to the total portfolio risk.
        """

        weights_df = pd.DataFrame(self.weights, index=self.asset_data.keys())

        stock_results = []
        for date in self.exposures.index:
            date_exposures = pd.DataFrame(
                [x.exposures.loc[date] for x in self.asset_data.values()],
                index=self.asset_data.keys(),
            )
            date_covar_matrix = pd.DataFrame(self.factor_covariance[date])
            stock_factor_risk = (date_exposures.dot(date_covar_matrix)).dot(
                date_exposures.T
            )
            weighted_stock_factor_risk = (
                weights_df.T.dot(stock_factor_risk) * weights_df.T
            )

            diag_specific_var_matrix = np.diag(
                [
                    x.risk_model_data.specific_risk[date] ** 2
                    for x in self.asset_data.values()
                ]
            )
            weighted_stock_specific_risk = (
                np.dot(weights_df.T, diag_specific_var_matrix) * weights_df.T
            )

            day_result = {'date': date}
            for stock, value in weighted_stock_factor_risk.loc[
                ZERO_IDX
            ].items():
                if value < 0:
                    value = -1 * (abs(value) ** 0.5)
                else:
                    value = value**0.5
                day_result.update({f'{stock}_factor': value})

            for stock, value in weighted_stock_specific_risk.loc[
                ZERO_IDX
            ].items():
                day_result.update({f'{stock}_specific': value**0.5})

            stock_results.append(day_result)

        stock_risks = pd.DataFrame(stock_results).set_index('date')

        stock_proportion_of_risk = self._get_proportion_of_risk(
            stock_risks, self.risk_model_data.total_risk
        )
        stock_contribution_to_risk = self._get_contribution_to_risk(
            stock_proportion_of_risk,
            self.risk_model_data.total_risk,
            annualised,
        )

        return stock_proportion_of_risk, stock_contribution_to_risk

    def get_factor_risk_by_stock(
        self, date: str, with_w: bool = False, annualised: bool = False
    ) -> pd.DataFrame:
        """
        Retrieves the factor risk contribution for each stock in the portfolio for a specific date.
        These are estimates for breaking down risk into each factor for each stock.

        Args:
            date (str): The specific date for which to retrieve factor risk data.
            with_w (bool): Whether to include the portfolio weights or not.
            annualised (bool, optional): Whether to annualize the risk values. Defaults to False.

        Returns:
            pd.DataFrame: A DataFrame containing factor risk values for each stock on the specified date.
        """
        results = {}
        specific_risk_t= np.empty(len(self.weights))
        i=0

        for asset, weight in zip(self.asset_data, self.weights):
            w = (weight if with_w else 1.0) * (np.sqrt(252) if annualised else 1)
            asset_risk = self.asset_data[asset].factor_risk.loc[date] * w
            results[asset] = asset_risk.to_dict()
            specific_risk_t[i] = self.asset_data[asset].risk_model_data.specific_risk[date]
            i=i+1

        total_risk_df= pd.DataFrame.from_dict(results).T
        total_risk_df = total_risk_df.assign(specific_risk = specific_risk_t)

        return total_risk_df

    def get_factor_attribution_by_stock_for_period(
        self, lookback: int = 1
    ) -> pd.DataFrame:
        """
        Calculates the return attribution from each factor for each stock in the portfolio over a lookback period,
        which represents the contribution of each factor to the weighted stock's return over the period.

        Args:
            lookback (int, optional): The number of days to look back for the attribution calculation. Defaults to 22.

        Returns:
            pd.DataFrame: A DataFrame containing the factor return attribution for each stock over the specified period.
        """
        result = {}
        for asset, weight in zip(self.asset_data, self.weights):
            factor_attribution = (
                self.asset_data[asset].get_factor_attribution()[-lookback:]
                * weight
            )
            period_attribution = (
                (1 + (factor_attribution / 100)).prod() - 1
            ) * 100
            result[asset] = period_attribution.to_dict()

        return pd.DataFrame.from_dict(result).T

    def get_weighted_stock_exposures_for_date(self, date: str) -> pd.DataFrame:
        """
        Retrieves the weighted factor exposures for each stock in the portfolio on a specific date.

        Args:
            date (str): The date for which to retrieve weighted exposures.

        Returns:
            pd.DataFrame: A DataFrame containing the weighted exposures for each stock in the portfolio on the given date.
        """
        result = pd.DataFrame(
            [
                x.exposures.loc[date] * w
                for x, w in zip(self.asset_data.values(), self.weights)
            ],
            index=self.asset_data.keys(),
        )
        return result

    def flatten_dict(self, d, parent_key=""):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}.{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self.flatten_dict(v, new_key).items())
            else:
                items.append((new_key, v))
        return dict(items)

    def get_total_working_days(self, start_date: str, end_date: str) -> int:
        # Generate a range of business days between the start and end dates
        business_days = pd.bdate_range(start=start_date, end=end_date)
        # Return the total number of business days
        return len(business_days)

    def remove_cash_positions(self, df_portfolio):
        return df_portfolio[
            ~df_portfolio.Identifier.str.contains('CASH')
        ].reset_index(drop=True)

    def check_missing_instrument_data(
        self,
        instrument,
        api_data,
        risk_model,
        date_from,
        date_to,
        working_days,
        df_portfolio_ex_missing,
    ):
        # Fetch risk model data for the instrument
        risk_model_data = api_data.get_risk_model_data(
            risk_model, instrument, date_from, date_to
        )

        # Check if data is insufficient or contains NaN values
        if (
            len(risk_model_data) < working_days
            or risk_model_data.isna().any().any()
        ):
            # Find the identifier and return it if data is missing or incomplete
            return df_portfolio_ex_missing[
                df_portfolio_ex_missing.Instrument == instrument
            ]['Identifier'].tolist()[0]
        return None  # Return None if the data is sufficient

    def get_portfolio_coverage(
        self,
        df_portfolio,
        risk_model_universe,
        risk_model,
        date_from,
        date_to,
        identifier_type,
    ):

        portfolio_universe = df_portfolio['Identifier'].astype(str).tolist()

        # SEDOLS need to have 7 characters. However, when importing from Excel,
        # the leading zeros are removed. This code adds them back.
        if identifier_type == 'SEDOL':
            portfolio_universe = [
                instrument.zfill(7) if len(instrument) < 7 else instrument
                for instrument in portfolio_universe
            ]

            df_portfolio['Identifier'] = portfolio_universe

        # Get identifiers that are not within the risk model universe.
        missing_identifiers = list(
            set(portfolio_universe) - set(risk_model_universe)
        )

        # Remove cash positions as we don't cover them.
        cash_identifiers = [
            identifier
            for identifier in portfolio_universe
            if 'CASH' in identifier
        ]

        missing_identifiers = missing_identifiers + cash_identifiers

        portfolio_ex_missing = [
            identifier
            for identifier in portfolio_universe
            if identifier not in missing_identifiers
        ]

        # Get instruments from portfolio identifiers.
        portfolio_mapping = api_instance.get_instruments_from_identifiers(
            {
                "identifiers": portfolio_ex_missing,
                "target_date": date_to,
            }
        )
        portfolio_identifiers = list(
            portfolio_mapping['resolved_instruments'].values()
        )

        # Get missing identifiers.
        missing_identifiers = (
            missing_identifiers + portfolio_mapping['unresolved_identifiers']
        )

        df_portfolio_ex_missing = df_portfolio[
            ~df_portfolio.Identifier.isin(missing_identifiers)
        ]
        df_portfolio_ex_missing.loc[:, 'Instrument'] = portfolio_identifiers

        # Check if any of the existing identifiers have missing historical data.
        api_data = ApiData()

        working_days = self.get_total_working_days(date_from, date_to)

        # Run the parallelized version
        results = Parallel(n_jobs=10)(
            delayed(self.check_missing_instrument_data)(
                instrument,
                api_data,
                risk_model,
                date_from,
                date_to,
                working_days,
                df_portfolio_ex_missing,
            )
            for instrument in portfolio_identifiers
        )

        # Filter out None values to get the list of instruments with missing historical data
        missing_historical_data = [
            result for result in results if result is not None
        ]

        covered_identifiers = [
            instrument
            for instrument in portfolio_identifiers
            if instrument
            not in df_portfolio_ex_missing[
                df_portfolio_ex_missing.Identifier.isin(
                    missing_historical_data
                )
            ]['Instrument'].tolist()
        ]

        return (
            missing_identifiers,
            missing_historical_data,
            covered_identifiers,
        )
